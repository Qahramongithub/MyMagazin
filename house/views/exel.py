import pandas as pd
from django.core.cache import cache
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from house.models import Product


@extend_schema(
    tags=["Excel"],
    request=None,
    responses={
        200: {
            'content': {'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': {}},
            'description': 'Excel file with products data'
        },
        404: {'description': 'Warehouse not found'}
    }
)
class ProductExcelExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        warehouse_id = cache.get(f"user_{user.id}_warehouse_id")

        if not warehouse_id:
            return HttpResponse("Warehouse not specified", status=404)

        qs = Product.objects.filter(
            warehouse_id=warehouse_id
        ).values('name', 'price', 'discount_price', 'base_price', 'quantity', 'min_quantity')

        df = pd.DataFrame(list(qs))

        if df.empty:
            df = pd.DataFrame(
                columns=['name', 'price', 'discount_price', 'base_price', 'quantity', 'investment', 'kassa',
                         'min_quantity']
            )
        else:
            df['discount_price'] = df['discount_price'].fillna(0)
            df['investment'] = df['quantity'] * df['base_price']
            df['kassa'] = df.apply(
                lambda r: r['quantity'] * (r['discount_price'] if r['discount_price'] > 0 else r['price']), axis=1
            )
        df = df.rename(columns={
            'name': 'Nomi',
            'price': 'Narxi',
            'discount_price': 'Chegirma narxi',
            'base_price': 'Bazaviy narx',
            'quantity': 'Miqdori',
            'min_quantity': 'Minimal miqdor',
            'investment': 'Xarajat',
            'kassa': 'Kassa'
        })

        # Uchta guruh
        df_above = df[df['quantity'] > df['min_quantity']]  # yashil
        df_between = df[(df['quantity'] > 0) & (df['quantity'] <= df['min_quantity'])]  # sariq
        df_zero = df[df['quantity'] == 0]  # qizil

        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # 1) quantity > min_quantity --> yashil sarlavha
        for row in dataframe_to_rows(df_above, index=False, header=True):
            ws.append(row)

        header_fill_green = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
        header_font_white = Font(color="FFFFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill_green
            cell.font = header_font_white

        # Bo‘sh qator
        ws.append([])
        ws.append([])

        # 2) 0 < quantity <= min_quantity --> sariq sarlavha
        start_row = ws.max_row + 1
        for r_idx, row in enumerate(dataframe_to_rows(df_between, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    cell.fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
                    cell.font = Font(bold=True)

        # Bo‘sh qator
        ws.append([])
        ws.append([])

        # 3) quantity == 0 --> qizil sarlavha
        start_row = ws.max_row + 1
        for r_idx, row in enumerate(dataframe_to_rows(df_zero, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    cell.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                    cell.font = Font(bold=True)

        # Ustun kengligi
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=products_warehouse_{warehouse_id}.xlsx'

        wb.save(response)
        return response
